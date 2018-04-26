# coding=utf-8
'''
    Python对Excel的读写主要有xlrd、xlwt、xlutils、openpyxl、xlsxwriter几种
    xlrd只能写xls文件格式
'''
import xlrd
import xlwt
import openpyxl

# 读取xlsx文件
xlsx = xlrd.open_workbook(u'file/枪V2.1版本数据指标1227.xlsx')
# 获取sheet列表
sheet_names = xlsx.sheet_names()

print('sheet名:',sheet_names[1])
# 指定sheet
sheet = xlsx.sheet_by_name(sheet_names[1])

print('共%s行，%s列' %(sheet.nrows,sheet.ncols))
for c in range(sheet.ncols):
    # 获取列内容
    cols = sheet.col_values(c)
    print('第%s列:' %(c+1),cols)

print('获取指定单元格内容：',sheet.cell(0,0))

# ------------------------------------------------------------------
f=xlwt.Workbook() #创建工作簿
sheet2 = f.add_sheet(u'新建sheet',cell_overwrite_ok=True) #创建sheet2
row0 = [u'姓名', u'年龄', u'出生日期', u'爱好', u'关系']
column0 = [u'小杰', u'小胖', u'小明', u'大神', u'大仙', u'小敏', u'无名']
# 生成第一行
for i in range(0, len(row0)):
    sheet2.write(0, i, row0[i])

# 生成第一列
for i in range(0, len(column0)):
    sheet2.write(i + 1, 0, column0[i])

sheet2.write(1, 2, '1991/11/11')
sheet2.write_merge(7, 7, 2, 4, u'暂无')  # 合并列单元格
sheet2.write_merge(1, 2, 4, 4, u'好朋友')  # 合并行单元格

f.save('file/demo1.xls')  # 保存文件 xlsx报错

# ------------------------------------------------------------------
workbook=openpyxl.Workbook()
booksheet = workbook.active     #获取当前活跃的sheet,默认是第一个sheet
booksheet.title = "New Shit"
#存第一行单元格cell(1,1)
booksheet.cell(coordinate=None, row=8, column=7).value = 666666   #这个方法索引从1开始
#存一行数据
workbook.save("file/test_openpyxl.xlsx")

# wb=xl.load_workbook('file/数据.xlsx')
# sheet1= wb.get_sheet_by_name('工作表1')
# sheet2= wb.get_sheet_by_name('工作表2')
#
# # x=sheet1.max_row
# x=100
#
# for i in range(1,x):
#     print(sheet1.cell(row=i, column=1).value,i)
#     # print(i)